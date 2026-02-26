import sqlite3
from openpyxl import load_workbook
from datetime import datetime
import os

DB_PATH = "home_maintenance.db"
EXCEL_PATH = r"E:\OneDrive\Home maintenance\Pool Maintenance\Swimming_Pool_Manager.xlsm"
SHEET_NAME = "Pool Test Log"

DATE_FMT = "%Y-%m-%d"


def ensure_pool_tests_table():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS pool_tests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        test_date TEXT NOT NULL,
        free_chlorine REAL NOT NULL,
        combined_chlorine REAL NOT NULL,
        total_chlorine REAL NOT NULL,
        salt_level REAL NOT NULL,
        alkalinity REAL NOT NULL,
        ph REAL NOT NULL,
        sunscreen REAL NOT NULL,
        hardness REAL NOT NULL,
        phosphates REAL NOT NULL,
        copper REAL NOT NULL,
        clarity_notes TEXT,
        actions_taken TEXT,
        next_test_date TEXT NOT NULL
    )
    """)

    conn.commit()
    conn.close()


def parse_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except:
        return None


def migrate_pool_tests():
    if not os.path.exists(EXCEL_PATH):
        print("Excel file not found:", EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet '{SHEET_NAME}' not found.")
        return

    ws = wb[SHEET_NAME]

    # Read header row
    headers = [c.value for c in ws[1]]

    # Expected headers
    expected = [
        "Date",
        "Free Chlorine (ppm)",
        "Combined Chlorine (ppm)",
        "Total Chlorine (ppm)",
        "Salt Level (ppm)",
        "Alkalinity (ppm)",
        "pH",
        "Sunscreen (Stabiliser) (ppm)",
        "Total Hardness (ppm)",
        "Phosphates (ppm)",
        "Copper Total (ppm)",
        "Water Clarity Notes",
        "Actions Taken (Chemicals Added, Adjustments, etc.)",
        "Next Planned Test Date",
    ]

    if headers[:len(expected)] != expected:
        print("ERROR: Excel headers do not match expected layout.")
        print("Found:", headers)
        return

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Optional: clear existing rows (comment out if you want to append)
    # cur.execute("DELETE FROM pool_tests")

    inserted = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # skip blank rows

        (
            date_val,
            fc,
            cc,
            tc,
            salt,
            alk,
            ph,
            sun,
            hard,
            phos,
            copper,
            clarity,
            actions,
            next_test,
        ) = row[:14]

        # Convert dates
        try:
            test_date = date_val.strftime(DATE_FMT)
        except:
            print("Skipping row with invalid date:", row)
            continue

        try:
            next_test_date = next_test.strftime(DATE_FMT)
        except:
            next_test_date = test_date  # fallback

        cur.execute("""
            INSERT INTO pool_tests (
                test_date,
                free_chlorine,
                combined_chlorine,
                total_chlorine,
                salt_level,
                alkalinity,
                ph,
                sunscreen,
                hardness,
                phosphates,
                copper,
                clarity_notes,
                actions_taken,
                next_test_date
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            test_date,
            parse_float(fc),
            parse_float(cc),
            parse_float(tc),
            parse_float(salt),
            parse_float(alk),
            parse_float(ph),
            parse_float(sun),
            parse_float(hard),
            parse_float(phos),
            parse_float(copper),
            clarity or "",
            actions or "",
            next_test_date,
        ))

        inserted += 1

    conn.commit()
    conn.close()

    print(f"Migration complete. Inserted {inserted} pool test records.")


def main():
    print("Starting Pool Test migration...")
    ensure_pool_tests_table()
    migrate_pool_tests()
    print("Done.")


if __name__ == "__main__":
    main()
